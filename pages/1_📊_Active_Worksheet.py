import streamlit as st
import pdfplumber
import pandas as pd
import re
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import random
import datetime
import math

st.set_page_config(page_title="Active Worksheet", layout="wide")

st.title("📊 Enterprise Engineering Control & Contract Worksheet")
st.write("This platform executes automated NEC compliance cross-examinations, structural voltage drop physics calculations, and advanced commercial contract breakouts.")

if "company_name" not in st.session_state:
    st.error("⚠️ Please return to the main Dashboard Gateway page to initialize your session parameters.")
    st.stop()

# Ensure global baseline variables are initialized safely
if "qty_journeymen" not in st.session_state: st.session_state.qty_journeymen = 1
if "rate_journeyman" not in st.session_state: st.session_state.rate_journeyman = 45.0
if "qty_helpers" not in st.session_state: st.session_state.qty_helpers = 1
if "rate_helper" not in st.session_state: st.session_state.rate_helper = 22.0
if "labor_burden_pct" not in st.session_state: st.session_state.labor_burden_pct = 0.30

total_crew_members = st.session_state.qty_journeymen + st.session_state.qty_helpers
raw_composite_rate = ((st.session_state.qty_journeymen * st.session_state.rate_journeyman) + (st.session_state.qty_helpers * st.session_state.rate_helper)) / total_crew_members
fully_burdened_labor_rate = raw_composite_rate * (1 + st.session_state.labor_burden_pct)

# --- DYNAMIC ENGINEERING SIMULATION OVERRIDES ---
st.write("### 🎛️ Engineering & System Design Constraints")
eng_col1, eng_col2, eng_col3 = st.columns(3)

with eng_col1:
    target_voltage = st.selectbox("Operating System Nominal Voltage", [120, 208, 240, 277, 480], index=0)
    circuit_amperage = st.number_input("Design Continuous Circuit Load (Amperes)", min_value=1.0, max_value=400.0, value=16.0, step=1.0)
    
with eng_col2:
    wire_size_selection = st.selectbox(
        "Design Wire Gauge Size (Copper)",
        options=["#14 AWG", "#12 AWG", "#10 AWG", "#8 AWG", "#6 AWG", "#4 AWG"],
        index=1
    )
    cm_map = {"#14 AWG": 4110, "#12 AWG": 6530, "#10 AWG": 10380, "#8 AWG": 16510, "#6 AWG": 26240, "#4 AWG": 41740}
    active_cm = cm_map[wire_size_selection]

with eng_col3:
    apply_kitting = st.checkbox("Enable Smart Assembly Kitting (Explode Counts)", value=True)

# --- PHOTOMETRIC & POWER DENSITY CALCULATION ENGINE ---
st.divider()
st.write("### 💡 Automated Photometric Layout & NEC Load Density Planner")

photo_col1, photo_col2, photo_col3 = st.columns(3)
with photo_col1:
    room_length = st.number_input("Target Area Room Length (Feet)", min_value=5.0, value=30.0, step=1.0)
    room_width = st.number_input("Target Area Room Width (Feet)", min_value=5.0, value=20.0, step=1.0)
    room_area = room_length * room_width
    st.write(f"Calculated Working Area: **{room_area:.1f} sq. ft.**")

with photo_col2:
    target_footcandles = st.slider("Target Illumination Intensity (Footcandles / LUX)", 10, 100, 40)
    fixture_lumen_output = st.number_input("Lumens Emitted per LED Fixture", min_value=500, max_value=10000, value=3200, step=100)

with photo_col3:
    occupancy_type = st.selectbox("NEC Space Classification (NEC Table 220.12)", ["Dwelling Unit / Residential (2.0 VA/sq.ft)", "Office Space (1.3 VA/sq.ft)", "Store / Retail (1.9 VA/sq.ft)"])
    va_multiplier = 2.0 if "Residential" in occupancy_type else (1.3 if "Office" in occupancy_type else 1.9)

# Compute Photometrics Matrix
cu_factor = 0.60
llf_factor = 0.85
calculated_fixtures_needed = math.ceil((target_footcandles * room_area) / (fixture_lumen_output * cu_factor * llf_factor))
nec_minimum_va_load = room_area * va_multiplier

p_res1, p_res2 = st.columns(2)
p_res1.success(f"💡 **Photometric Recommendation:** Deploy **{calculated_fixtures_needed} Recessed LED Fixtures** in this zone.")
p_res2.info(f"⚡ **NEC Load Calculation:** Minimum continuous power footprint: **{nec_minimum_va_load:,.1f} Volt-Amperes (VA)**.")

# --- READ DISTRIBUTOR MATRIX RATES ---
if "vendor_pricing" not in st.session_state:
    st.session_state.vendor_pricing = {
        "Main Panel Enclosure": 450.00, "GFCI Receptacle": 18.00, 
        "Disconnect Switch": 85.00, "Single Pole Switch": 1.50,
        "3/4\" EMT Conduit Run (Linear Ft)": 1.25
    }

conduit_base_cost = st.session_state.vendor_pricing.get("3/4\" EMT Conduit Run (Linear Ft)", 1.25)
price_ratio = conduit_base_cost / 1.25

# --- AGGREGATE SPATIAL HISTORY ---
total_raw_footage = 0.0
active_zone_tag = "General Branch Run"

if "sheet_ledger" in st.session_state:
    for sheet_id, data in st.session_state.sheet_ledger.items():
        total_raw_footage += data["conduit_runs"] + data["vertical_drops"]
        if data["conduit_runs"] > 0:
            active_zone_tag = data["active_zone"]

rows_to_compile = []

if calculated_fixtures_needed > 0:
    rows_to_compile.append({
        "Item Name": "Specification Grade 2x2 Recessed LED Troffer", "Phase": "Trim-Out", "Zone/Location": "Photometric Layout Zone",
        "Detected Qty": int(calculated_fixtures_needed), "Unit Cost ($)": 65.00, "Mins to Install": 25
    })

if total_raw_footage > 0:
    conduit_sticks = math.ceil(total_raw_footage / 10.0)
    rows_to_compile.append({
        "Item Name": f"3/4\" EMT Conduit (10ft Sticks) - Formed for {wire_size_selection}", "Phase": "Rough-In", "Zone/Location": active_zone_tag,
        "Detected Qty": int(conduit_sticks), "Unit Cost ($)": round(6.50 * price_ratio, 2), "Mins to Install": 12
    })
    
    couplings_needed = max(conduit_sticks - 1, 0)
    if couplings_needed > 0:
        rows_to_compile.append({
            "Item Name": "3/4\" EMT Set-Screw Coupling", "Phase": "Rough-In", "Zone/Location": active_zone_tag,
            "Detected Qty": int(couplings_needed), "Unit Cost ($)": round(1.15 * price_ratio, 2), "Mins to Install": 3
        })
        
    straps_needed = math.ceil(total_raw_footage / 8.0) + 2
    rows_to_compile.append({
        "Item Name": "3/4\" 1-Hole EMT Strap (NEC 358.30)", "Phase": "Rough-In", "Zone/Location": active_zone_tag,
        "Detected Qty": int(straps_needed), "Unit Cost ($)": round(0.45 * price_ratio, 2), "Mins to Install": 2
    })

# --- PROCESS BLUEPRINT REGEX DATA TEXT SCANS ---
with st.sidebar:
    st.header("🔍 Custom Regex Profiles")
    panel_kw = st.text_input("Main Panel Keywords", value="panel, load center, mlo")
    gfci_kw = st.text_input("GFCI Keywords", value="gfci, gfi, ground fault")
    disc_kw = st.text_input("Disconnect Keywords", value="disconnect, safety switch")
    switch_kw = st.text_input("Switch Keywords", value="single pole, 1-pole switch")

@st.cache_data
def process_and_scan_blueprint(uploaded_file_bytes, p_kw, g_kw, d_kw, s_kw, prices_dict):
    pdf_file = BytesIO(uploaded_file_bytes)
    full_text = ""
    with pdfplumber.open(pdf_file) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if text: full_text += text + "\n"
            
    def clean_pattern(raw_input):
        parts = [p.strip() for p in raw_input.split(",")]
        return r"(\d+)\s*(?:-|x)?\s*(?:amp)?\s*(?:" + "|".join(parts) + ")"

    electrical_manifest = {
        "Main Panel Enclosure": {"pattern": clean_pattern(p_kw), "cost": prices_dict["Main Panel Enclosure"], "mins": 120, "phase": "Rough-In", "zone": "Service Room"},
        "GFCI Receptacle": {"pattern": clean_pattern(g_kw), "cost": prices_dict["GFCI Receptacle"], "mins": 20, "phase": "Trim-Out", "zone": "Wet Areas (Kitchen/Bath)"},
        "Disconnect Switch": {"pattern": clean_pattern(d_kw), "cost": prices_dict["Disconnect Switch"], "mins": 45, "phase": "Rough-In", "zone": "HVAC / Equipment"},
        "Single Pole Switch": {"pattern": clean_pattern(s_kw), "cost": prices_dict["Single Pole Switch"], "mins": 15, "phase": "Trim-Out", "zone": "General Lighting"}
    }
    
    scanned_results = []
    for item, info in electrical_manifest.items():
        matches = re.findall(info["pattern"], full_text, re.IGNORECASE)
        total_qty = sum(int(match) for match in matches if match.isdigit())
        scanned_results.append({
            "Item Name": item, "Phase": info["phase"], "Zone/Location": info["zone"],
            "Detected Qty": total_qty, "Unit Cost ($)": info["cost"], "Mins to Install": info["mins"]
        })
    return scanned_results

if st.session_state.uploaded_file_bytes is not None:
    base_data = process_and_scan_blueprint(st.session_state.uploaded_file_bytes, panel_kw, gfci_kw, disc_kw, switch_kw, st.session_state.vendor_pricing)
else:
    base_data = []

# --- SMART ASSEMBLY KITTING EXPLODER MATRIX ---
final_compiled_rows = []
for row in base_data:
    item = row["Item Name"]
    qty = row["Detected Qty"]
    
    if qty > 0 and apply_kitting:
        if item == "GFCI Receptacle":
            final_compiled_rows.append({"Item Name": "Commercial Grade 20A GFCI Device", "Phase": "Trim-Out", "Zone/Location": row["Zone/Location"], "Detected Qty": qty, "Unit Cost ($)": row["Unit Cost ($)"], "Mins to Install": 12})
            final_compiled_rows.append({"Item Name": "4\" Square Deep Steel Box (2-1/8\")", "Phase": "Rough-In", "Zone/Location": row["Zone/Location"], "Detected Qty": qty, "Unit Cost ($)": 3.10, "Mins to Install": 6})
            final_compiled_rows.append({"Item Name": "1-Gang Devia Plaster Mud Ring (1/2\")", "Phase": "Rough-In", "Zone/Location": row["Zone/Location"], "Detected Qty": qty, "Unit Cost ($)": 1.85, "Mins to Install": 3})
            final_compiled_rows.append({"Item Name": "Stainless Steel Single-Gang Faceplate", "Phase": "Trim-Out", "Zone/Location": row["Zone/Location"], "Detected Qty": qty, "Unit Cost ($)": 1.45, "Mins to Install": 2})
            final_compiled_rows.append({"Item Name": "#10-32 Green Grounding Pigtailed Clip", "Phase": "Rough-In", "Zone/Location": row["Zone/Location"], "Detected Qty": qty * 2, "Unit Cost ($)": 0.35, "Mins to Install": 1})
        elif item == "Single Pole Switch":
            final_compiled_rows.append({"Item Name": "Specification Grade 20A Toggle Switch", "Phase": "Trim-Out", "Zone/Location": row["Zone/Location"], "Detected Qty": qty, "Unit Cost ($)": row["Unit Cost ($)"], "Mins to Install": 10})
            final_compiled_rows.append({"Item Name": "Handy Utility Metal Wall Box", "Phase": "Rough-In", "Zone/Location": row["Zone/Location"], "Detected Qty": qty, "Unit Cost ($)": 2.25, "Mins to Install": 5})
            final_compiled_rows.append({"Item Name": "Industrial Toggle Switch Wall Plate", "Phase": "Trim-Out", "Zone/Location": row["Zone/Location"], "Detected Qty": qty, "Unit Cost ($)": 0.95, "Mins to Install": 2})
        else:
            final_compiled_rows.append(row)
    else:
        final_compiled_rows.append(row)

master_df = pd.DataFrame(final_compiled_rows)
if rows_to_compile:
    canvas_df = pd.DataFrame(rows_to_compile)
    master_df = pd.concat([master_df, canvas_df], ignore_index=True)

# Inject Vision Computer Vision rows
for item_name, count in st.session_state.vision_counts.items():
    if count > 0:
        vision_row = pd.DataFrame([{
            "Item Name": f"AI Scan: {item_name}", "Phase": "Trim-Out", "Zone/Location": "Vision Takeoff Area", 
            "Detected Qty": count, "Unit Cost ($)": 24.50, "Mins to Install": 25
        }])
        master_df = pd.concat([master_df, vision_row], ignore_index=True)

st.divider()
st.write("### 📋 Formulated Bill of Materials & Subcontractor Proposal")
edited_df = st.data_editor(master_df, num_rows="dynamic", use_container_width=True)

edited_df["Detected Qty"] = pd.to_numeric(edited_df["Detected Qty"]).fillna(0)
edited_df["Unit Cost ($)"] = pd.to_numeric(edited_df["Unit Cost ($)"]).fillna(0)
edited_df["Mins to Install"] = pd.to_numeric(edited_df["Mins to Install"]).fillna(0)

# Execute Financial Aggregations
edited_df["Line_Material"] = edited_df["Detected Qty"] * edited_df["Unit Cost ($)"]
edited_df["Line_Labor"] = (edited_df["Detected Qty"] * edited_df["Mins to Install"] / 60) * fully_burdened_labor_rate

total_mat = edited_df["Line_Material"].sum()
total_labor = edited_df["Line_Labor"].sum()
base_contract_value = (total_mat + total_labor) * (1 + st.session_state.overhead)

# --- NEW PILLAR 1: SCHEDULE OF VALUES (SOV) MILESTONE GENERATOR ---
st.divider()
st.write("### 📑 Schedule of Values (SOV) Progressive Project Breakdown")
st.caption("Commercial project breakdown showing itemized project costs mapped cleanly to specialized project installation phases.")

sov_rows = []
for phase, group in edited_df.groupby("Phase"):
    phase_mat = group["Line_Material"].sum()
    phase_labor = group["Line_Labor"].sum()
    phase_total_with_markup = (phase_mat + phase_labor) * (1 + st.session_state.overhead)
    allocation_pct = (phase_total_with_markup / base_contract_value * 100) if base_contract_value > 0 else 0
    
    sov_rows.append({
        "Construction Phase Milestone": phase,
        "Allocated Material Cost": round(phase_mat, 2),
        "Allocated Burdened Labor": round(phase_labor, 2),
        "Total Milestone Billing Sum": round(phase_total_with_markup, 2),
        "Contract Allocation Ratio": f"{allocation_pct:.1f}%"
    })

sov_df = pd.DataFrame(sov_rows)
st.dataframe(sov_df, use_container_width=True, hide_index=True)

# --- NEW PILLAR 2: DYNAMIC CHANGE-ORDER VARIATION ENGINE ---
st.divider()
st.write("### 📝 Active Project Change-Order Management System")
st.caption("Layer field variations over the baseline contract pricing profile to calculate adjusted net values dynamically.")

with st.expander("➕ Open Live Change-Order Modification Console", expanded=False):
    co_col1, co_col2, co_col3 = st.columns(3)
    with co_col1:
        co_title = st.text_input("Change Order Designation Title", value="CO #1: Owner Added Kitchen Circuits")
    with co_col2:
        co_mat_adjust = st.number_input("Variation Material Adjustment ($ Value)", value=0.0, step=50.0)
    with co_col3:
        co_hours_adjust = st.number_input("Variation Labor Adjustment (Man-Hours)", value=0.0, step=1.0)
        
    co_labor_cost = co_hours_adjust * fully_burdened_labor_rate
    co_total_gross = (co_mat_adjust + co_labor_cost) * (1 + st.session_state.overhead)
    adjusted_final_contract = base_contract_value + co_total_gross
    
    st.write("---")
    m_col1, m_col2, m_col3 = st.columns(3)
    m_col1.metric("Original Baseline Contract Value", f"${base_contract_value:,.2f}")
    m_col2.metric("Net Change-Order Cost Impact", f"${co_total_gross:,.2f}", delta=f"{co_hours_adjust:.1f} Hrs Variant Impact")
    m_col3.metric("Adjusted Target Contract Value", f"${adjusted_final_contract:,.2f}", delta="Dynamic Variation Sync Active")

st.divider()
st.write("### 📥 Executive Proposal Distribution")

# --- HIGH-FIDELITY EXCEL COMPILATION HOOK ---
def generate_executive_excel(df_data, mat_cost, labor_cost, total_bid, overhead_pct, rate, comp_name, sov_dataframe):
    output = BytesIO()
    wb = openpyxl.Workbook()
    font_family = "Segoe UI"
    charcoal_fill = PatternFill(start_color="262626", end_color="262626", fill_type="solid")
    gold_accent_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    ice_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    title_font = Font(name=font_family, size=18, bold=True, color="262626")
    section_font = Font(name=font_family, size=12, bold=True, color="FFFFFF")
    bold_font = Font(name=font_family, size=11, bold=True)
    regular_font = Font(name=font_family, size=11)
    thin_border = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'), top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))
    
    ws1 = wb.active; ws1.title = "Executive Summary"; ws1.views.sheetView[0].showGridLines = True
    ws1["A1"] = f"{comp_name.upper()} PROPOSAL"; ws1["A1"].font = title_font
    ws1.merge_cells("A4:B4"); ws1["A4"] = "PROJECT FINANCIAL SUMMARY"; ws1["A4"].fill = charcoal_fill; ws1["A4"].font = section_font
    
    metrics = [("Material Cost Subtotal", mat_cost), ("Burdened Labor Allocation", labor_cost), ("Blended Crew Composite Rate ($/hr)", rate), ("Markup Burden Percentage", overhead_pct)]
    for idx, (m, v) in enumerate(metrics, start=5):
        ws1[f"A{idx}"] = m; ws1[f"A{idx}"].font = regular_font; ws1[f"A{idx}"].border = thin_border
        ws1[f"B{idx}"] = v; ws1[f"B{idx}"].font = regular_font; ws1[f"B{idx}"].border = thin_border
        ws1[f"B{idx}"].number_format = '0.0%' if idx == 8 else '$#,##0.00'
        
    ws1["A10"] = "TOTAL TARGET CONTRACT BID"; ws1["A10"].font = bold_font; ws1["A10"].fill = gold_accent_fill; ws1["A10"].border = thin_border
    ws1["B10"] = total_bid; ws1["B10"].font = Font(name=font_family, size=12, bold=True, color="C00000"); ws1["B10"].fill = gold_accent_fill; ws1["B10"].border = thin_border; ws1["B10"].number_format = '$#,##0.00'
    
    # Inject SOV Table directly into Excel layout
    start_row = 13
    ws1.cell(row=start_row, column=1, value="SCHEDULE OF VALUES (SOV) PHASE BREAKOUT").font = section_font
    ws1.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=4)
    ws1.cell(row=start_row, column=1).fill = charcoal_fill
    
    sov_headers = ["Phase", "Material Cost", "Burdened Labor", "Total Milestone Sum"]
    for c_idx, h in enumerate(sov_headers, start=1):
        cell = ws1.cell(row=start_row+1, column=c_idx, value=h)
        cell.font = bold_font; cell.fill = ice_fill; cell.border = thin_border
        
    for r_idx, r_data in sov_dataframe.iterrows():
        cur_r = start_row + 2 + r_idx
        ws1.cell(row=cur_r, column=1, value=r_data["Construction Phase Milestone"]).font = regular_font
        ws1.cell(row=cur_r, column=2, value=r_data["Allocated Material Cost"]).number_format = '$#,##0.00'
        ws1.cell(row=cur_r, column=3, value=r_data["Allocated Burdened Labor"]).number_format = '$#,##0.00'
        ws1.cell(row=cur_r, column=4, value=r_data["Total Milestone Billing Sum"]).number_format = '$#,##0.00'
        for c in range(1, 5): ws1.cell(row=cur_r, column=c).border = thin_border
    
    ws2 = wb.create_sheet(title="Bill of Materials"); ws2.views.sheetView[0].showGridLines = True
    headers = ["Component Name", "Phase", "Target Zone", "Takeoff Qty", "Estimated Unit Cost"]
    for idx, t in enumerate(headers, start=1):
        cell = ws2.cell(row=1, column=idx, value=t); cell.fill = charcoal_fill; cell.font = section_font
        
    for r_idx, r_data in df_data.iterrows():
        r = r_idx + 2
        ws2.cell(row=r, column=1, value=r_data["Item Name"]).font = regular_font
        ws2.cell(row=r, column=2, value=r_data["Phase"]).font = regular_font
        ws2.cell(row=r, column=3, value=r_data["Zone/Location"]).font = regular_font
        ws2.cell(row=r, column=4, value=r_data["Detected Qty"]).number_format = '#,##0'
        ws2.cell(row=r, column=5, value=r_data["Unit Cost ($)"]).number_format = '$#,##0.00'
        if r_idx % 2 == 0:
            for col_c in range(1, 6): ws2.cell(row=r, column=col_c).fill = ice_fill
        for col_c in range(1, 6): ws2.cell(row=r, column=col_c).border = thin_border

    for sheet in [ws1, ws2]:
        for col in sheet.columns:
            vals = [str(cell.value or '') for cell in col]
            max_len = max(len(v) for v in vals) if vals else 10
            col_letter = get_column_letter(col[0].column)
            sheet.column_dimensions[col_letter].width = max(max_len + 3, 14)

    wb.save(output)
    return output.getvalue()

excel_data = generate_executive_excel(edited_df, total_mat, total_labor, base_contract_value, st.session_state.overhead, fully_burdened_labor_rate, st.session_state.company_name, sov_df)
st.download_button("🚀 Export Executive Proposal Package (.xlsx)", data=excel_data, file_name="Executive_Bid.xlsx")